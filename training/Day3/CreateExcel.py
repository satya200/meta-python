
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

ws["C5"] = "Hello World"
ws["D5"] = 5000
ws["C7"] = datetime.now()

wb.save("myFirstExcel.xlsx")
