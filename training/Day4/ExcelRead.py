
from openpyxl import load_workbook

wb = load_workbook("firstExcel.xlsx")

wb.active = wb['Employees']
ws = wb.active

mySheet = wb["Sheet"]
mySheet.title = "Comcast"

print("-" * 40)
print(wb.sheetnames)

print(ws.dimensions)
print("-" * 40)

dataRange = ws[ws.dimensions]
print(dataRange)

print("-" * 40)

for c1, c2, c3, c4, c5 in dataRange:
    print("{0:5}\t{1:5}\t{2:5}\t{3:5}\t{4:5}".format(c1.value, c2.value,
                                                     c3.value, c4.value, c5.value))

wb.save("firstExcel.xlsx")