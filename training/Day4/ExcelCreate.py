"""
Workbook  => Excel file

"""

from openpyxl import Workbook

wb = Workbook()
wb.create_sheet("Employees")
wb.active = wb['Employees']
print(wb.sheetnames)

wb['Sheet'].active = False
ws = wb.active

cell = ws['A4']

data = [
    ['empid', 'empname', 'age', 'dept', 'salary'],
    [501, 'Jack', 28, "FIN", 45000],
    [250, 'Jill', 27, 'HR', 35000],
    [105, 'Mike', 32, 'TL', 68500]
]

for x in data:
    ws.append(x)

wb.save("firstExcel.xlsx")