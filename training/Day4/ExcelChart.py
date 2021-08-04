
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("firstExcel.xlsx")

wb.create_sheet("Chart")

wb['Employees'].active = False
wb.active = wb['Chart']

ws = wb.active

data = (
    ('Products', "Sales"),
    ('Pepsi', 250),
    ('Coke', 180),
    ('Sprite', 165),
    ('thumbs up', 280),
    ('mirinda', 200)
)

for row in data:
    ws.append(row)

dataRange = ws.dimensions

print("-" * 40)

cellRef = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=6)
labelRef = Reference(ws, min_col=1, min_row=2, max_row=6)

chart = BarChart()
chart.add_data(cellRef)
chart.set_categories(labelRef)
chart.title = "Beverage Sales"
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in lakhs"


ws.add_chart(chart, "E6")
wb.save("firstExcel.xlsx")